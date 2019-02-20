#! /usr/bin/python
# encoding:utf-8

#读写xls
#import xlrd
#import xlwt
# 读写xlsx
import openpyxl
from sklearn import preprocessing
from sklearn.model_selection import train_test_split
import numpy as np
from scipy import sparse
from imblearn.combine import SMOTEENN
from hive_client import HiveClient
import pickle
import os
from dj_log import log

class PreprocModule():
	'''
	指标预处理模块
	'''
	def __init__(self, save_standard_d_path=None, save_standard_c_path=None):
		'''
		初始化方法
		Args:
			save_standard_d_path: 标准化样本离散属性方法存放路径
			save_standard_c_path: 标准化样本连续属性方法存放路径
		'''
		self._save_standard_d_path = save_standard_d_path
		self._save_standard_c_path = save_standard_c_path
	
	def write2xls():
		'''
		将数据写入.xls文件
		'''
		# TO DO

	def write2xlsx(self, data, save_path):
		'''
		将数据写入.xlsx文件
		Args:
			data: 待写入数据
			save_path: 保存路径
		Returns:
			None
		'''
		wb = openpyxl.Workbook()
		ws = wb.active
		for row in data:
			ws.append(row)
			wb.save(save_path)
  
	def readxls():
		'''
		从xls文件中读数据
		'''
		#TO DO

	def readxlsx(self, read_path, sheet_name):
		'''
		从xlsx文件中读数据
		Args:
			read_path: 读入数据路径
			sheet_name: Excel表sheet名称
		Returns:
			X: 读入的数据特征 type -> 'list'
			y: 读入的数据标记 type -> 'list'
		'''
		wb = openpyxl.load_workbook(read_path)
		sheet = wb[sheet_name]
		rows = sheet.max_row
		cols = sheet.max_column
		ori_data = [[0 for j in range(cols)] for i in range(rows) ]
		for row in sheet.rows:
			for cell in row:
				ori_data[cell.row-1][cell.col_idx-1] = cell.value
		ori_data = ori_data[1:] #第一列为标题头
		X, y = [item[4:-1] for item in ori_data if item[-1] != "0"], [int(item[-1]) for item in ori_data if item[-1] != "0"] #前4个指标为展示指标,不参与训练
		X = [[int(float(value)) for value in item] for item in X]
		return X, y
  
	def readHive(self, hql):
		'''
		从线上Hive表中读取数据
		'''
		hive_c = HiveClient()
		status, ori_data = hive_c.query(hql)
		if status == 0:
			X, y = [item[4:-1] for item in ori_data if item[-1] != "0"], [int(item[-1]) for item in ori_data if item[-1] != "0"] #前4个指标为展示指标,不参与训练
			X = [[int(float(value)) for value in item] for item in X]
			return X, y
		log.error("Query data from Hive failed.")
		return
  
	def posRateStatis(self, y):
		'''
		统计样本中正样例占比
		Args:
			y: 数据样本标记 type -> 'list'
		Returns:
			p_rate: 正样本占比 range -> [0,1]
		'''
		Y_p, Y_n = y.count(1), y.count(-1)
		p_rate = np.sum(Y_p) / float(len(y)) #小于20%时应该进行样本不平衡处理
		return p_rate
		
	def standardization(self, X, enc_d=None, enc_c=None):
		'''
		One-hot Encoding(仅针对离散指标)
		Min-Max-Score标准化(仅针对数值指标)
		Args:
			X: 数据特征样本
			enc_d: 离散属性特征标准化方法
			enc_c: 连续属性特征标准化方法
		Returns:
			X: 标准化后的数据特征样本
			d_cols: 离散指标编码后的新特征数
			enc_d: 离散属性特征标准化方法
			enc_c: 连续属性特征标准化方法
		'''
		X_d, X_c = [item[:15] for item in X], [item[15:] for item in X]
		if enc_d is None:
			enc = preprocessing.OneHotEncoder()
			enc_d = enc.fit(X_d)
			if self._save_standard_d_path is not None:
				#if not os.path.exists(self._save_standard_d_path):
				#	os.makedirs(self._save_standard_d_path)
				pickle.dump(enc_d, open(self._save_standard_d_path,'wb'))
				log.debug("Save standardization discrete features method success. The path is %s", self._save_standard_d_path)
		if enc_c is None:
			min_max_scaler = preprocessing.MinMaxScaler()
			enc_c = min_max_scaler.fit(X_c)
			if self._save_standard_c_path is not None:
				#if not os.path.exists(self._save_standard_c_path):
				#	os.makedirs(self._save_standard_c_path)
				pickle.dump(enc_c, open(self._save_standard_c_path,'wb'))
				log.debug("Save standardization continuous features method success. The path is %s", self._save_standard_c_path)
		X_enc_d = enc_d.transform(X_d)
		X_sca_c = enc_c.transform(X_c)
		_, d_cols = X_enc_d.shape #离散指标One-hot Encoding后的列数(新指标数)
		#print('discrete: ', X_enc_d.shape)
		#print('continuous: ', X_sca_c.shape)
		X = sparse.hstack((X_enc_d, X_sca_c)).toarray()
		return X, d_cols, enc_d, enc_c
	  
	def imbalanceProcess(self, X, y):
		'''
		样本不平衡处理
		Args:
			X: 待处理的数据特征样本
			y: 待处理的数据标记样本
		Returns:
			X: 处理后的数据特征样本
			y: 处理后的数据标记样本
		'''
		sm = SMOTEENN()
		X, y = sm.fit_sample(X, y)
		return X, y